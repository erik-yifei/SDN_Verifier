import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Specify the paths to the files
client_file_path = r'C:\Users\ErikWang\Documents\SDN_Verifier\Files\OFAC_SDN_List_09.12.2024.xlsx' # Change date in this line to reflect the most recent update
alt_file_path = r'C:\Users\ErikWang\Documents\SDN_Verifier\Files\alt.csv'
add_file_path = r'C:\Users\ErikWang\Documents\SDN_Verifier\Files\add.csv'
sdn_file_path = r'C:\Users\ErikWang\Documents\SDN_Verifier\Files\sdn.csv'

# Load the client Excel file
df = pd.read_excel(client_file_path, engine='openpyxl')

# Specify the columns to use
columns_to_use = [1, 2]  # Assuming column B is 'Last Name' and column C is 'First Name'

# Load only the specified columns from the client Excel file
clients_df = pd.read_excel(client_file_path, usecols=columns_to_use, engine='openpyxl')
clients_df.columns = ['Last', 'First']  # Rename columns for clarity

# Combine the 'Last Name' and 'First Name' columns into a single 'Full Name' column
clients_df['Full'] = clients_df['Last'] + ', ' + clients_df['First']

# Load the SDN CSV files
alt_df = pd.read_csv(alt_file_path, header=None)
add_df = pd.read_csv(add_file_path, header=None)
sdn_df = pd.read_csv(sdn_file_path, header=None)

# Check for overlaps between client names and SDN lists for persons
overlaps_alt_person = clients_df[clients_df['Full'].isin(alt_df[3])]
overlaps_add_person = clients_df[clients_df['Full'].isin(add_df[2])]
overlaps_sdn_person = clients_df[clients_df['Full'].isin(sdn_df[1])]

# Check for overlaps between client names and SDN lists for corporations
overlaps_alt_corp = clients_df[clients_df['Last'].isin(alt_df[3])]
overlaps_add_corp = clients_df[clients_df['Last'].isin(add_df[2])]
overlaps_sdn_corp = clients_df[clients_df['Last'].isin(sdn_df[1])]

# Remove rows where 'First' is NaN to avoid incorrectly combined names
overlaps_alt_person = overlaps_alt_person.dropna(subset=['First'])
overlaps_add_person = overlaps_add_person.dropna(subset=['First'])
overlaps_sdn_person = overlaps_sdn_person.dropna(subset=['First'])

# Combine results and remove duplicates
all_overlaps_person = pd.concat([overlaps_alt_person, overlaps_add_person, overlaps_sdn_person]).drop_duplicates()
all_overlaps_corp = pd.concat([overlaps_alt_corp, overlaps_add_corp, overlaps_sdn_corp]).drop_duplicates()

# Print the overlaps
print("Overlapping names for ALT (Persons):")
print(overlaps_alt_person)
print("Overlapping names for ADD (Persons):")
print(overlaps_add_person)
print("Overlapping names for SDN (Persons):")
print(overlaps_sdn_person)

print("Overlapping names for ALT (Corporations):")
print(overlaps_alt_corp)
print("Overlapping names for ADD (Corporations):")
print(overlaps_add_corp)
print("Overlapping names for SDN (Corporations):")
print(overlaps_sdn_corp)

# Create a set of all matched names
matched_names_person = set(all_overlaps_person['Full'])
matched_names_corp = set(all_overlaps_corp['Last'])

# Define a function to check if there is a complete match
def check_match(row):
    if row['Full'] in matched_names_person or row['Last'] in matched_names_corp:
        return 'Match'
    else:
        return 'No Match'

# Apply the function to create the 'SDN Status' column (complete match)
clients_df['SDN Status'] = clients_df.apply(check_match, axis=1)

# Set the first two rows to blank for 'SDN Status'
clients_df.loc[:1, 'SDN Status'] = ''

# Print the updated clients_df with the 'SDN Status' column
print(clients_df)

# Save the updated clients_df to a temporary Excel file to retain formatting
temp_file_path = r'C:\Users\ErikWang\Documents\SDN_Verifier\Files\client_match_status.xlsx'
clients_df.to_excel(temp_file_path, index=False)

# Load the workbook and select the active worksheet
workbook = load_workbook(temp_file_path)
worksheet = workbook.active

# Define the fill styles for matches
match_fill = PatternFill(start_color="D30000", end_color="D30000", fill_type="solid")  # Red

# Apply the fill styles based on the 'SDN Status' column
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    sdn_status = row[worksheet.max_column - 1].value  # Adjusted to match the correct column

    if sdn_status == 'Match':
        for cell in row:
            cell.fill = match_fill

# Save the workbook with highlights
workbook.save(temp_file_path)

